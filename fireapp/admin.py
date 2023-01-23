from django.contrib.auth.hashers import make_password
from django.db.models.fields import NOT_PROVIDED
from import_export import widgets
import tablib
import traceback
from django.db.transaction import (
    TransactionManagementError,
    savepoint,
    savepoint_commit,
    savepoint_rollback,
)
from django.core.exceptions import (
    FieldDoesNotExist,
    ImproperlyConfigured,
    ValidationError,
)
from copy import deepcopy
from import_export.results import Error, Result, RowResult
from django.template.response import TemplateResponse
from django.utils.translation import gettext_lazy as _
from django.contrib import admin
from django.contrib.auth.admin import UserAdmin
from django.apps import apps
from django.http import HttpResponse, HttpResponseRedirect
from django.urls import reverse
from import_export.fields import Field
from import_export.widgets import ForeignKeyWidget, Widget
from fireapp.models import (
    ADMIN_TYPE,
    CONST_TYPE_ADMIN,
    CONST_TYPE_STUDENT,
    CONST_TYPE_TEACHER,
    STUDENT_TYPE,
    TEACHER_TYPE,
    Course,
    Department,
    Grade,
    Section,
    Subject,
    Student,
    Teacher,
    Quiz,
    QuizData,
)
from django.contrib.auth import get_user_model
from .forms import CustomUserChangeForm, CustomUserCreateBaseForm
from import_export import resources
from import_export.admin import ImportExportModelAdmin, ImportMixin
from django.core.exceptions import ObjectDoesNotExist
from openpyxl import Workbook, load_workbook
import openpyxl
from django.contrib.auth import get_user_model
from django.core.exceptions import PermissionDenied
from django.views.decorators.http import require_POST
from django.utils.encoding import force_str
from django.utils.decorators import method_decorator
from io import BytesIO
import logging
logger = logging.getLogger(__name__)
# from tablib.formats import registry
# from tablib.formats._xlsx import XLSXFormat


# class FormattedXLSX(XLSXFormat):
#     @classmethod
#     def export_set(cls, dataset, freeze_panes=True, formatter=None):
#         """Returns XLSX representation of Dataset."""
#         wb = Workbook()
#         ws = wb.worksheets[0]
#         ws.title = dataset.title if dataset.title else "Tablib Dataset"

#         cls.dset_sheet(dataset, ws, freeze_panes=freeze_panes)

#         ### Just added this lines to original code
#         if formatter:
#             formatter(ws)
#         ###

#         stream = BytesIO()
#         wb.save(stream)
#         return stream.getvalue()


# registry.register("xlsx", FormattedXLSX())


CustomUser = get_user_model()
GRADE_WORKSHEET = "RAW GRADES"
SEMESTRAL_WORKSHEET = "SEMESTRAL GRADE"
EXCEL_ROW_START_OFFSET = 10
EXCEL_COL_START_OFFSET = 2
SCHOOL_YEARS_SEMESTER = ("", "FIRST", "SECOND", "THIRD", "FOURTH", "FIFTH")
STUDENTS_WORKSHEET = "Students"
EXEMPTED_MODELS = (Quiz, QuizData)


@admin.register(Course)
class CourseAdmin(admin.ModelAdmin):
    model = Course
    list_display = (
        "id",
        "course_abbr",
        "course_name",
        "get_department_name",
        "course_credit",
    )
    list_editable = ["course_credit"]
    filter_horizontal = ("subjects", )

    @admin.display(description="Department Name", ordering="department__name")
    def get_department_name(self, obj):
        if obj.course_department is not None:
            return obj.course_department.name
        else:
            return ''

    def save_related(self, request, form, formsets, change):
        obj = form.instance
        print(request)
        print(form)
        obj.save()
        super().save_related(request, form, formsets, change)


class StudentInline(admin.StackedInline):
    model = Student
    can_delete = False


@admin.register(Teacher)
class CustomTeacherAdmin(admin.ModelAdmin):
    model = Teacher
    list_display = (
        "get_teacher_name",
        "teacher_id",
        "department",
    )
    readonly_fields = ("user", "get_teacher_name")
    filter_horizontal = ("sections", "subjects")

    sortable_by = ("teacher_id" "department",)

    fieldsets = (
        (
            None,
            {
                "fields": (
                    "user",
                    "get_teacher_name",
                    "teacher_id",
                    "department",
                    "sections",
                    "subjects",
                )
            },
        ),
    )

    def get_teacher_name(self, obj):
        return obj.user.first_name + " " + obj.user.last_name

    get_teacher_name.admin_order_field = "Teacher"
    get_teacher_name.short_description = "Teacher Name"

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.user_type == CONST_TYPE_TEACHER:
            return qs.filter(user=request.user)
        else:
            return qs

    def changelist_view(self, request, extra_context=None):
        if request.user.user_type == CONST_TYPE_TEACHER:
            user = request.user
            return HttpResponseRedirect(
                reverse(
                    "admin:%s_%s_change"
                    % (self.model._meta.app_label, self.model._meta.model_name),
                    args=(user.id,),
                )
            )
        return super().changelist_view(request=request, extra_context=extra_context)





@admin.register(Student)
class CustomStudentAdmin(admin.ModelAdmin):

    model = Student
    list_display = (
        "get_student_name",
        "scholar_no",
        "section",
        "course",
        "current_year",
        "current_semester",
    )
    list_filter = ("section",)
    search_fields = ("scholar_no",)
    readonly_fields = ("student_email", "get_student_name")

    sortable_by = (
        "scholar_no",
        "section",
        "current_year",
        "current_semester",
    )

    fieldsets = (
        (
            None,
            {
                "fields": (
                    "scholar_no",
                    "student_email",
                    "get_student_name",
                    "section",
                    "course",
                    "current_year",
                    "current_semester",
                    "mobile",
                    "parents_mobile",
                    "home_address",
                )
            },
        ),
    )

    def student_email(self, obj):
        return obj.user.email

    def get_student_name(self, obj):
        return obj.user.first_name + " " + obj.user.last_name

    get_student_name.admin_order_field = "Student"  # Allows column order sorting
    get_student_name.short_description = "Student Name"  # Renames column head

    def changelist_view(self, request, extra_context=None):
        if request.user.user_type == CONST_TYPE_STUDENT:
            user = request.user
            return HttpResponseRedirect(
                reverse(
                    "admin:%s_%s_change"
                    % (self.model._meta.app_label, self.model._meta.model_name),
                    args=(user.id,),
                )
            )
        return super().changelist_view(request=request, extra_context=extra_context)

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.user_type == CONST_TYPE_STUDENT:
            return qs.filter(user=request.user)
        else:
            return qs

    def get_readonly_fields(self, request, obj=None):
        if obj and request.user.user_type == CONST_TYPE_STUDENT:
            # We are adding an object
            return self.readonly_fields + ("section", "scholar_no")
        else:
            return self.readonly_fields

    def has_add_permission(self, request):
        return False

class MyForeignKeyWidget(ForeignKeyWidget):
    def get_queryset(self, value, row, *args, **kwargs):
        return self.model.objects.filter(
            first_name__iexact=row["first_name"],
            last_name__iexact=row["last_name"]
        )

class MyField(Field):
    def save(self, obj, data, is_m2m=False, **kwargs):
        """
        If this field is not declared readonly, the object's attribute will
        be set to the value returned by :meth:`~import_export.fields.Field.clean`.
        """
        if not self.readonly:
            attrs = self.attribute.split('__')
            for attr in attrs[:-1]:
                obj = getattr(obj, attr, None)
            cleaned = self.clean(data, **kwargs)
            if cleaned is not None or self.saves_null_values:
                if not is_m2m:
                    setattr(obj, attrs[-1], cleaned)
                else:
                    getattr(obj, attrs[-1]).set(cleaned)

    def clean(self, data, **kwargs):
        """
        Translates the value stored in the imported datasource to an
        appropriate Python object and returns it.
        """
        try:
            value = data[self.column_name]
        except KeyError:
            raise KeyError("Column '%s' not found in dataset. Available "
                           "columns are: %s" % (self.column_name, list(data)))

        # If ValueError is raised here, import_obj() will handle it
        value = self.widget.clean(value, row=data, **kwargs)

        if value in self.empty_values and self.default != NOT_PROVIDED:
            if callable(self.default):
                return self.default()
            return self.default

        return value


class MyWidget(ForeignKeyWidget):
    def clean(self, value, row=None, *args, **kwargs):
        # val = super().clean(value)
        if value:
            if self.field == 'section_name':
                section = Section.objects.filter(section_name=value).first()
                return section
            if self.field == 'course_name':
                course = Course.objects.filter(course_name=value).first()
                return course# self.get_queryset(value, row, *args, **kwargs).get(**{self.field: val})
        else:
            return None


class CustomUserResource(resources.ModelResource):
    email = Field(attribute='email', column_name='email')
    gender = Field(attribute='gender', column_name='gender')
    birth_date = Field(attribute='birth_date', column_name='birth_date')
    first_name = Field(attribute='first_name', column_name='first_name')
    middle_name = Field(attribute='middle_name', column_name='middle_name')
    last_name = Field(attribute='last_name', column_name='last_name')
    password = Field(attribute='raw_password', column_name='password')
    section = Field(attribute='section_name', column_name='section')
    course = Field(attribute='course_name', column_name='course')
    scholar_no = Field(attribute='uid', column_name='scholar_no')
    date_enrolled = Field(attribute='date_enrolled')
    # section = Field()
    # course = Field()
    current_year = Field(attribute='current_year')
    current_semester = Field(attribute='current_semester')
    mobile = Field(attribute='mobile')
    parents_mobile = Field(attribute='parents_mobile')
    home_address = Field(attribute='home_address')

    section = MyField(
                    column_name='section',
                    attribute='section',
                    widget=MyWidget(Section, 'section_name'))

    course = MyField(
        column_name='course',
        attribute='course',
        widget=MyWidget(Course, 'course_name'))

    def __init__(self):
        # The fields class attribute is the *class-wide* definition of
        # fields. Because a particular *instance* of the class might want to
        # alter self.fields, we create self.fields here by copying cls.fields.
        # Instances should always modify self.fields; they should not modify
        # cls.fields.
        self.fields = deepcopy(self.fields)

        # lists to hold model instances in memory when bulk operations are enabled
        self.create_instances = list()
        self.update_instances = list()
        self.delete_instances = list()

        self.student_instances = {}

    class Meta:
        worksheet_title = STUDENTS_WORKSHEET
        model = CustomUser
        # import_id_fields = ('id', )
        exclude = (
            "uid",
        )
        fields = ('id', 'scholar_no', 'email', 'gender', 'birth_date',
                  'first_name', 'middle_name', 'last_name',
                  'password', 'date_enrolled',
                  'section', 'course',
                  'current_year', 'current_semester', 'mobile',
                  'parents_mobile', 'home_address')

        export_order = ('id', 'scholar_no', 'email', 'gender', 'birth_date',
                        'first_name', 'middle_name', 'last_name',
                        'password', 'date_enrolled',
                        'section', 'course',
                        'current_year', 'current_semester', 'mobile',
                        'parents_mobile', 'home_address')
        skip_unchanged = True
        report_skipped = True

    def dehydrate_date_enrolled(self, obj):
        return self.get_column_value(obj, 'date_enrolled')

    def dehydrate_section(self, obj):
        return self.get_column_value(obj, 'section')

    def dehydrate_course(self, obj):
        return self.get_column_value(obj, 'course')

    def dehydrate_current_year(self, obj):
        return self.get_column_value(obj, 'current_year')

    def dehydrate_current_semester(self, obj):
        return self.get_column_value(obj, 'current_semester')

    def dehydrate_mobile(self, obj):
        return self.get_column_value(obj, 'mobile')

    def dehydrate_parents_mobile(self, obj):
        return self.get_column_value(obj, 'parents_mobile')

    def dehydrate_home_address(self, obj):
        return self.get_column_value(obj, 'home_address')

    def dehydrate_scholar_no(self, obj):
        return self.get_column_value(obj, 'scholar_no')

    def get_column_value(self, obj, key):
        if not obj.uid:
            return ''
        if obj.uid not in self.student_instances:
            self.student_instances[obj.uid] = Student.objects.filter(user=obj).first()
        
        val = getattr(obj, key, '')
        if not val:
            val = str(getattr(self.student_instances[obj.uid], key, ''))
            if isinstance(val, NOT_PROVIDED):
                val = ''
        return val
            

    def after_import(self, dataset, result, using_transactions, dry_run, **kwargs):
        """
        Override to add additional logic. Does nothing by default.
        """
        a = 1

    def after_import_instance(self, instance, new, row_number=None, **kwargs):
        """
        Override to add additional logic. Does nothing by default.
        """
        # if instance.uid not in self.student_instances:
        #     self.student_instances[instance.uid] = Student.objects.filter(user=instance).first()

        
        pass

    def import_obj(self, obj, data, dry_run, **kwargs):
        """
        Traverses every field in this Resource and calls
        :meth:`~import_export.resources.Resource.import_field`. If
        ``import_field()`` results in a ``ValueError`` being raised for
        one of more fields, those errors are captured and reraised as a single,
        multi-field ValidationError."""
        errors = {}
        for field in self.get_import_fields():
            if isinstance(field.widget, widgets.ManyToManyWidget):
                continue
            try:
                self.import_field(field, obj, data, **kwargs)
            except Section.DoesNotExist:
                pass
            except Course.DoesNotExist:
                pass
            except ValueError as e:
                errors[field.attribute] = ValidationError(
                    force_str(e), code="invalid")
        if errors:
            raise ValidationError(errors)

    def import_field(self, field, obj, data, is_m2m=False, **kwargs):
        """
        Calls :meth:`import_export.fields.Field.save` if ``Field.attribute``
        is specified, and ``Field.column_name`` is found in ``data``.
        """
        if field.attribute and field.column_name in data:
            field.save(obj, data, is_m2m, **kwargs)

    def import_row(self, row, instance_loader, using_transactions=True, dry_run=False, raise_errors=False, **kwargs):
        """
        Imports data from ``tablib.Dataset``. Refer to :doc:`import_workflow`
        for a more complete description of the whole import process.

        :param row: A ``dict`` of the row to import

        :param instance_loader: The instance loader to be used to load the row

        :param using_transactions: If ``using_transactions`` is set, a transaction
            is being used to wrap the import

        :param dry_run: If ``dry_run`` is set, or error occurs, transaction
            will be rolled back.
        """
        skip_diff = self._meta.skip_diff
        row_result = self.get_row_result_class()()
        original = None
        try:
            self.before_import_row(row, **kwargs)
            instance, new = self.get_or_init_instance(instance_loader, row)
            self.after_import_instance(instance, new, **kwargs)
            if new:
                row_result.import_type = RowResult.IMPORT_TYPE_NEW
            else:
                row_result.import_type = RowResult.IMPORT_TYPE_UPDATE
            row_result.new_record = new
            if not skip_diff:
                original = deepcopy(instance)
                diff = self.get_diff_class()(self, original, new)
            if self.for_delete(row, instance):
                if new:
                    row_result.import_type = RowResult.IMPORT_TYPE_SKIP
                    if not skip_diff:
                        diff.compare_with(self, None, dry_run)
                else:
                    row_result.import_type = RowResult.IMPORT_TYPE_DELETE
                    row_result.add_instance_info(instance)
                    self.delete_instance(instance, using_transactions, dry_run)
                    if not skip_diff:
                        diff.compare_with(self, None, dry_run)
            else:
                import_validation_errors = {}
                try:
                    self.import_obj(instance, row, dry_run, **kwargs)
                except ValidationError as e:
                    # Validation errors from import_obj() are passed on to
                    # validate_instance(), where they can be combined with model
                    # instance validation errors if necessary
                    import_validation_errors = e.update_error_dict(import_validation_errors)
                if self.skip_row(instance, original):
                    row_result.import_type = RowResult.IMPORT_TYPE_SKIP
                else:
                    self.validate_instance(instance, import_validation_errors)
                    self.save_instance(instance, using_transactions, dry_run)
                    self.save_m2m(instance, row, using_transactions, dry_run)
                    row_result.add_instance_info(instance)
                if not skip_diff:
                    diff.compare_with(self, instance, dry_run)

            if not skip_diff and not self._meta.skip_html_diff:
                row_result.diff = diff.as_html()
            self.after_import_row(row, row_result, **kwargs)

        except ValidationError as e:
            row_result.import_type = RowResult.IMPORT_TYPE_INVALID
            row_result.validation_error = e
        except Exception as e:
            row_result.import_type = RowResult.IMPORT_TYPE_ERROR
            # There is no point logging a transaction error for each row
            # when only the original error is likely to be relevant
            if not isinstance(e, TransactionManagementError):
                logger.debug(e, exc_info=e)
            tb_info = traceback.format_exc()
            row_result.errors.append(self.get_error_result_class()(e, tb_info, row))

        if self._meta.use_bulk:
            # persist a batch of rows
            # because this is a batch, any exceptions are logged and not associated
            # with a specific row
            if len(self.create_instances) == self._meta.batch_size:
                self.bulk_create(using_transactions, dry_run, raise_errors, batch_size=self._meta.batch_size)
            if len(self.update_instances) == self._meta.batch_size:
                self.bulk_update(using_transactions, dry_run, raise_errors, batch_size=self._meta.batch_size)
            if len(self.delete_instances) == self._meta.batch_size:
                self.bulk_delete(using_transactions, dry_run, raise_errors)

        if dry_run is False:
            if instance.password is None:
                instance.password = make_password(instance.raw_password)
            instance.is_active = True
            instance.is_staff = True
            instance.save()
            for key, value in row.items():
                if instance.student:
                    if key == 'section':
                        value = Section.objects.filter(section_name=value).first()
                    if key == 'course':
                        value = Course.objects.filter(course_name=value).first()
                    # value = getattr(self.student_instances[instance.uid], key, '')
                    setattr(instance.student, key, value)

            instance.student.save()

        return row_result
    
    # def before_import_row(self, row, row_number=None, **kwargs):
    #     """
    #     Override to add additional logic. Does nothing by default.
    #     """
    #     section = Section.objects.filter(section_name=row['section']).first()
    #     if section is not None:
    #         row['section'] = section.id

    #     course = Course.objects.filter(course_name=row['course']).first()
    #     if course is not None:
    #         row['course'] = course.id

    def export(self, queryset=None, *args, **kwargs):
        """
        Exports a resource.
        """

        self.before_export(queryset, *args, **kwargs)

        if queryset is None:
            queryset = self.get_queryset()
        headers = self.get_export_headers()
        data = tablib.Dataset(headers=headers)

        for obj in self.iter_queryset(queryset):
            if obj.user_type == CONST_TYPE_STUDENT:
                data.append(self.export_resource(obj))

        self.after_export(queryset, data, *args, **kwargs)

        return data

    def export_field(self, field, obj):
        field_name = self.get_field_name(field)
        method = getattr(self, 'dehydrate_%s' % field_name, None)
        if method is not None:
            return method(obj)
        return field.export(obj)


@admin.register(CustomUser)
class CustomUserAdmin(ImportExportModelAdmin, UserAdmin):

    def get_export_data(self, file_format, queryset, *args, **kwargs):
        """
        Returns file_format representation for given queryset.
        """
        request = kwargs.pop("request")
        if not self.has_export_permission(request):
            raise PermissionDenied

        data = self.get_data_for_export(request, queryset, *args, **kwargs)
        if self.resource_class and self.resource_class._meta.worksheet_title:
            data.title = self.resource_class._meta.worksheet_title
        export_data = file_format.export_data(data)
        encoding = kwargs.get("encoding")
        if not file_format.is_binary() and encoding:
            export_data = export_data.encode(encoding)
        return export_data



    # @method_decorator(require_POST)
    # def process_import(self, request, *args, **kwargs):
    #     """
    #     Perform the actual import action (after the user has confirmed the import)
    #     """
    #     if not self.has_import_permission(request):
    #         raise PermissionDenied

    #     form_type = self.get_confirm_import_form()
    #     confirm_form = form_type(request.POST)
    #     if confirm_form.is_valid():
    #         import_formats = self.get_import_formats()
    #         input_format = import_formats[
    #             int(confirm_form.cleaned_data["input_format"])
    #         ]()
    #         tmp_storage = self.get_tmp_storage_class()(
    #             name=confirm_form.cleaned_data["import_file_name"]
    #         )
    #         data = tmp_storage.read(input_format.get_read_mode())
    #         if not input_format.is_binary() and self.from_encoding:
    #             data = force_str(data, self.from_encoding)
    #         dataset = input_format.create_dataset(data)

    #         workbook = load_workbook(
    #             filename=BytesIO(data), read_only=True, data_only=True
    #         )

    #         self.process_excel(workbook)

    #         result = self.process_dataset(
    #             dataset, confirm_form, request, *args, **kwargs
    #         )

    #         tmp_storage.remove()

    #         return self.process_result(result, request)

    # def process_excel(self, workbook):
    #     students_worksheet = workbook.get_sheet_by_name(STUDENTS_WORKSHEET)
        
    #     data = []
    #     for row in students_worksheet.iter_rows():  # Offset for header
            
    #         scholar_no = None
    #         try:
    #             scholar_no = row[0].value
    #         except Exception:
    #             pass
    #         if scholar_no is None or not scholar_no:
    #             continue

    #         student = Student.objects.filter(scholar_no=scholar_no).first()
    #         if student is None:
    #             User = get_user_model()
    #             user = User.objects.create_student(uid=scholar_no, password=scholar_no)
    #             if user is None:
    #                 continue
    #             full_name = row[1].value.strip().split(",")
    #             user.last_name = full_name[0]
    #             middle_initial = full_name[1][-2] if full_name[1][-1] == "." else ""
    #             user.first_name = full_name[1].replace(f"{middle_initial}.", "").strip()
    #             user.middle_name = middle_initial
    #             user.save()
    #             student = Student.objects.get(user=user)

    #         if student is None:
    #             continue

    #         grade.student = student
    #         grade.student.scholar_no = scholar_no

    #         grade.student.save()
    #         grade.student_year = year
    #         grade.student_semester = semester
    #         try:
    #             average = row[113].value
    #             grade.average = round(average, 2)
    #             grade.grade = row[114].value
    #         except IndexError:
    #             pass
    #         data.append(grade)

    #     # Bulk create data
    #     Grade.objects.bulk_create(data)

    resource_class = CustomUserResource
    add_form = CustomUserCreateBaseForm
    form = CustomUserChangeForm
    ordering = ("email",)
    list_display = (
        "uid",
        "email",
        "first_name",
        "last_name",
        "_user_type",
        "is_active",
    )
    list_filter = (
        "user_type",
        "is_active",
    )
    search_fields = ("uid", "email", "first_name", "last_name")
    filter_horizontal = (
        "groups",
        "user_permissions",
    )
    # inlines = []
    exclude = ("is_superuser", "last_login", "date_joined")
    readonly_fields = ("uid", "_user_type",)

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.user_type == CONST_TYPE_TEACHER:
            return qs.exclude(user_type=CONST_TYPE_ADMIN)
        elif request.user.user_type == CONST_TYPE_STUDENT:
            return qs.filter(id=request.user.id)
        else:
            return qs

    fieldsets = (
        ("User Information", {"fields": ("email", "uid","_user_type")}),
        (
            "Personal Information",
            {
                "fields": (
                    "first_name",
                    "last_name",
                    "middle_name",
                    "birth_date",
                    "gender",
                )
            },
        ),
        ("Permissions", {"fields": ("is_active", "groups", "user_permissions")}),
    )

    add_fieldsets = (
        (
            "User Information",
            {
                "classes": ("wide",),
                "fields": (
                    "uid",
                    "email",
                    "password1",
                    "password2",
                    "user_type",
                    "is_active",
                ),
            },
        ),
        (
            "Personal Information",
            {
                "classes": ("wide",),
                "fields": (
                    "first_name",
                    "last_name",
                    "middle_name",
                    "birth_date",
                    "gender",
                ),
            },
        ),
    )

    def has_add_permission(self, request):
        if (request.user.user_type == CONST_TYPE_ADMIN or
                request.user.user_type == CONST_TYPE_TEACHER):
            return True
        else:
            return False

    def _user_type(self, obj):
        return dict(ADMIN_TYPE + TEACHER_TYPE + STUDENT_TYPE)[int(obj.user_type)]

    def changelist_view(self, request, extra_context=None):
        if request.user.user_type == CONST_TYPE_STUDENT:
            user = request.user
            return HttpResponseRedirect(
                reverse(
                    "admin:%s_%s_change"
                    % (self.model._meta.app_label, self.model._meta.model_name),
                    args=(user.id,),
                )
            )
        return super().changelist_view(request=request, extra_context=extra_context)

    def get_readonly_fields(self, request, obj=None):
        if obj and not request.user.user_type == CONST_TYPE_ADMIN:
            # We are adding an object
            return self.readonly_fields + ("user_permissions", "groups", "is_active")
        else:
            return self.readonly_fields

    def formfield_for_choice_field(self, db_field, request, **kwargs):
        if db_field.name == "user_type":
            if request.user.user_type == CONST_TYPE_ADMIN:
                kwargs["choices"] = ADMIN_TYPE + TEACHER_TYPE
            if request.user.user_type == CONST_TYPE_TEACHER:
                kwargs["choices"] = STUDENT_TYPE

        return super().formfield_for_choice_field(db_field, request, **kwargs)

    def save_model(self, request, obj, form, change):
        obj.is_superuser = obj.user_type == CONST_TYPE_ADMIN
        obj.is_staff = 1
        obj.save()

    
    


@admin.register(Subject)
class SubjectAdmin(admin.ModelAdmin):
    list_display = ("subject_name", "unit")
    list_filter = ("subject_name", "unit")


@admin.register(Department)
class DepartmentAdmin(admin.ModelAdmin):
    readonly_fields = [
        "teachers",
    ]

    def teachers(self, obj):
        # this_section = Section.objects.get(id=object_id)
        this_teachers = Teacher.objects.filter(department=obj)
        result = list()
        for teacher in this_teachers:
            result.append(
                f"{teacher.teacher_id} - {teacher.user.last_name} \
                {teacher.user.first_name}"
            )
        return "\n".join(result)

    teachers.admin_order_field = "Teachers"  # Allows column order sorting
    teachers.short_description = "Teachers"  # Renames column head


@admin.register(Section)
class SectionAdmin(admin.ModelAdmin):

    readonly_fields = [
        "students",
    ]

    def students(self, obj):
        # this_section = Section.objects.get(id=object_id)
        this_students = Student.objects.filter(section=obj)
        result = list()
        for student in this_students:
            result.append(
                f"{student.scholar_no} - {student.user.last_name} \
                {student.user.first_name}"
            )
        return "\n".join(result)

    students.admin_order_field = "Students"  # Allows column order sorting
    students.short_description = "Students"  # Renames column head


class GradeResource(resources.ModelResource):
    class Meta:
        model = Grade
        fields = (
            "id",
            "grade",
            "student",
        )
        skip_unchanged = True
        report_skipped = False


@admin.register(Grade)
class GradeAdmin(ImportMixin, admin.ModelAdmin):
    resource_class = GradeResource
    list_display = (
        "student",
        "grade",
        "subject",
        "student_year",
        "student_semester",
    )
    exclude = ('average', )
    list_filter = (
        "student",
        "subject",
        "student_year",
        "student_semester",
    )

    def get_list_filter(self, request):
        list_filter = list(self.list_filter)
        if request.user.user_type == CONST_TYPE_STUDENT:
            list_filter.remove("student")

        return list_filter

    def get_list_display(self, request):
        """ Removes the E column unless "Yes" has been selected in the 
            dummy filter.
        """
        list_display = list(self.list_display)
        if request.user.user_type == CONST_TYPE_STUDENT:
            list_display.remove("student")

        return list_display

    def has_import_permission(self, request):
        if request.user.user_type == CONST_TYPE_STUDENT:
            return False
        return True

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.user_type == CONST_TYPE_STUDENT:
            this_student = Student.objects.get(user=request.user)
            return qs.filter(student=this_student)
        else:
            return qs

    def get_readonly_fields(self, request, obj=None):
        if obj and request.user.user_type == CONST_TYPE_STUDENT:
            return self.readonly_fields + ("subject",)
        else:
            return self.readonly_fields

    @method_decorator(require_POST)
    def process_import(self, request, *args, **kwargs):
        """
        Perform the actual import action (after the user has confirmed the import)
        """
        if not self.has_import_permission(request):
            raise PermissionDenied

        form_type = self.get_confirm_import_form()
        confirm_form = form_type(request.POST)
        if confirm_form.is_valid():
            import_formats = self.get_import_formats()
            input_format = import_formats[
                int(confirm_form.cleaned_data["input_format"])
            ]()
            tmp_storage = self.get_tmp_storage_class()(
                name=confirm_form.cleaned_data["import_file_name"]
            )
            data = tmp_storage.read(input_format.get_read_mode())
            if not input_format.is_binary() and self.from_encoding:
                data = force_str(data, self.from_encoding)
            dataset = input_format.create_dataset(data)

            workbook = load_workbook(
                filename=BytesIO(data), read_only=True, data_only=True
            )

            self.process_excel(workbook)

            result = self.process_dataset(
                dataset, confirm_form, request, *args, **kwargs
            )

            tmp_storage.remove()

            return self.process_result(result, request)

    def process_excel(self, workbook):
        semestral_worksheet = workbook.get_sheet_by_name(SEMESTRAL_WORKSHEET)
        grades_worksheet = workbook.get_sheet_by_name(GRADE_WORKSHEET)

        year = 1
        semester = 1
        try:
            school_year = str(semestral_worksheet.cell(row=17, column=3).value).upper()
            year = SCHOOL_YEARS_SEMESTER.index(school_year.split(" ")[0])
            school_semester = semestral_worksheet.cell(row=19, column=3).value
            semester = SCHOOL_YEARS_SEMESTER.index(school_semester.split(" ")[0])
        except IndexError:
            pass

        data = []

        for row in grades_worksheet.iter_rows(
            min_row=EXCEL_ROW_START_OFFSET, min_col=EXCEL_COL_START_OFFSET
        ):  # Offset for header
            grade = Grade()
            scholar_no = None
            try:
                scholar_no = row[0].value
            except Exception:
                pass
            if scholar_no is None or not scholar_no:
                continue

            student = Student.objects.filter(scholar_no=scholar_no).first()
            if student is None:
                User = get_user_model()
                user = User.objects.create_student(uid=scholar_no, password=scholar_no)
                if user is None:
                    continue
                full_name = row[1].value.strip().split(",")
                user.last_name = full_name[0]
                middle_initial = full_name[1][-2] if full_name[1][-1] == "." else ""
                user.first_name = full_name[1].replace(f"{middle_initial}.", "").strip()
                user.middle_name = middle_initial
                user.save()
                student = Student.objects.get(user=user)

            if student is None:
                continue

            grade.student = student
            grade.student.scholar_no = scholar_no

            grade.student.save()
            grade.student_year = year
            grade.student_semester = semester
            try:
                average = row[113].value
                grade.average = round(average, 2)
                grade.grade = row[114].value
            except IndexError:
                pass
            data.append(grade)

        # Bulk create data
        Grade.objects.bulk_create(data)


app_config = apps.get_app_config("fireapp")
models = app_config.get_models()

for model in models:
    if model.__class__ in EXEMPTED_MODELS:
        continue
    try:
        admin.site.register(model)
    except admin.sites.AlreadyRegistered:
        pass

admin.site.unregister(Quiz)
admin.site.unregister(QuizData)


